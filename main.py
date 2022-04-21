from openpyxl import Workbook
import sys

def main():
    global item_id
    item_id = 1
    height = 0
    width = 0
    exit_flag = 0
    price = int(input("请输入单价："))#単価を入力してください
    while exit_flag == 0:
        try:
            print("输入任意字母退出，输入999重新输入")#任意のアルファベットを入力すると終了し、999を入力すると再入力できます。
            height = int(input("请输入长："))#長さを入力してください
            if height == 999: continue
            width = int(input("请输入宽："))#幅を入力してください
            if width == 999: continue
            area = int(height) * int(width)
            total = area * price
            sheet.append([item_id, height, width, area, price, total])
            item_id = int(sheet.cell(row=sheet.max_row, column=1).value) + 1
        except Exception as e:
            print(e)
            exit_flag = 1

def sumOftotal():
    rows = sheet.rows
    sum = 0
    for row in range(2, sheet.max_row+1):
        print(row)
        sum += int(sheet.cell(row =row, column=6).value)
    sheet.cell(row=sheet.max_row+1, column=sheet.max_column).value = "合计金额"#合計金額
    sheet.cell(row=sheet.max_row+1, column=sheet.max_column).value = sum


if __name__ == "__main__":
    wb = Workbook()
    sheet = wb.active
    sheet.title = "预算报价"
    sheet.append(["id", "长", "宽", "面积", "单价", "总价"])#長さ、幅、面積、単価、合計
    main()
    sumOftotal()
    wb.save("预算报价.xlsx")#見積書を生成する
    sys.exit()
